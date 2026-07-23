from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List
from collections import defaultdict
import pymysql, pymysql.cursors, os, csv, io, json, base64
from datetime import date, datetime
from zoneinfo import ZoneInfo
import numpy as np
import cv2
import httpx

IST = ZoneInfo("Asia/Kolkata")
from dotenv import load_dotenv

load_dotenv()

# Setup Local uploads folder
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BACKEND_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def save_base64_file(base64_str: str, filename: str) -> str:
    """Saves a base64 string locally to the uploads directory."""
    try:
        if "," in base64_str:
            header, base64_str = base64_str.split(",")
        
        file_data = base64.b64decode(base64_str)
        file_path = os.path.join(UPLOAD_DIR, filename)
        # Ensure the parent directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, "wb") as f:
            f.write(file_data)
        url_path = filename.replace("\\", "/")
        return f"/uploads/{url_path}"
    except Exception as e:
        print(f"[File Save Error] {e}")
        if base64_str.startswith("http"):
            return base64_str
        raise e

# Initialize local FaceAnalysis offline
HF_URL = os.getenv("HF_URL", "https://aravind-20-arcface-api.hf.space")
face_app = None
try:
    from insightface.app import FaceAnalysis
    # Use CPU provider for local single-image inference stability
    face_app = FaceAnalysis(name='buffalo_l', providers=['CPUExecutionProvider'])
    face_app.prepare(ctx_id=-1, det_size=(640, 640))
    print("SUCCESS: Local FaceAnalysis initialized successfully (using insightface on CPU)")
except Exception as e:
    print(f"WARNING: Local FaceAnalysis could not be initialized: {e}")

def local_get_embedding(image_base64: str) -> dict:
    if face_app is None:
        return {"success": False, "error": "Face analysis model not initialized"}
    try:
        if "," in image_base64:
            image_base64 = image_base64.split(",")[1]
        img_bytes = base64.b64decode(image_base64)
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            return {"success": False, "error": "Invalid image format"}
        faces = face_app.get(img)
        if not faces:
            return {"success": False, "error": "No face detected"}
        face = max(faces, key=lambda f: (f.bbox[2]-f.bbox[0]) * (f.bbox[3]-f.bbox[1]))
        return {
            "success": True,
            "embedding": face.normed_embedding.tolist(),
            "bbox": face.bbox.tolist()
        }
    except Exception as e:
        print(f"[Local ArcFace Error] {e}")
        return {"success": False, "error": str(e)}

def hf_get_embedding(image_base64: str) -> dict:
    try:
        if "," in image_base64:
            image_base64 = image_base64.split(",")[1]
        res = httpx.post(
            f"{HF_URL}/get-embedding",
            json={"image": image_base64},
            timeout=60,
        )
        data = res.json()
        emb = data.get("embedding") or data.get("embeddings") or data
        bbox = data.get("bbox")
        return {"success": True, "embedding": emb, "bbox": bbox}
    except Exception as e:
        print(f"[HF Error] get-embedding: {e}")
        return {"success": False, "error": str(e)}

app = FastAPI(title="AttendTrack API")
app.add_middleware(CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET","POST","PUT","DELETE","OPTIONS"],
    allow_headers=["*"])

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

def get_db():
    ssl = {"ssl": {"ssl_mode": "REQUIRED"}} if os.getenv("DB_SSL", "true").lower() == "true" else {}
    return pymysql.connect(
        host=os.getenv("DB_HOST"),
        port=int(os.getenv("DB_PORT","4000")),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME","attendtrack"),
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
        **ssl
    )

@app.on_event("startup")
def init_tables():
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS employees (
            id              INT AUTO_INCREMENT PRIMARY KEY,
            full_name       VARCHAR(150) NOT NULL,
            father_name     VARCHAR(150),
            email           VARCHAR(255) UNIQUE NOT NULL,
            phone           VARCHAR(15),
            aadhaar_no      VARCHAR(12)  UNIQUE NOT NULL,
            department      VARCHAR(100) NOT NULL,
            location        VARCHAR(200),
            source          VARCHAR(150),
            shift_hrs       DECIMAL(4,1) DEFAULT 9.0,
            face_descriptor  JSON,
            face_image       VARCHAR(500),
            aadhaar_pdf      VARCHAR(500),
            account_name     VARCHAR(150),
            account_number   VARCHAR(30),
            ifsc             VARCHAR(15),
            pan              VARCHAR(12),
            project_name     VARCHAR(150),
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS attendance (
            id        INT AUTO_INCREMENT PRIMARY KEY,
            emp_id    INT NOT NULL,
            date      DATE NOT NULL,
            log_in  TIME,
            log_out TIME,
            total_hrs DECIMAL(5,2) DEFAULT 0,
            ot_hrs    DECIMAL(5,2) DEFAULT 0,
            status    ENUM('on-duty','present','absent') DEFAULT 'absent',
            UNIQUE KEY unique_emp_date (emp_id, date),
            FOREIGN KEY (emp_id) REFERENCES employees(id) ON DELETE CASCADE)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS source_persons (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            name       VARCHAR(150) NOT NULL UNIQUE,
            account_name   VARCHAR(150),
            account_number VARCHAR(30),
            ifsc           VARCHAR(15),
            pan            VARCHAR(12),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS admin_settings (
            id               INT AUTO_INCREMENT PRIMARY KEY,
            pay_per_day      DECIMAL(10,2) DEFAULT 500.00,
            ot_pay_per_hr    DECIMAL(10,2) DEFAULT 100.00,
            food_allowance   DECIMAL(10,2) DEFAULT 50.00,
            food_before_time VARCHAR(5)    DEFAULT '08:00',
            updated_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)""")

        # Upgrade existing schema — add missing columns if not present
        for col, defn in [
            ("food_allowance",   "DECIMAL(10,2) DEFAULT 120.00"),
            ("food_before_time", "VARCHAR(5)    DEFAULT '08:00'"),
            ("father_name",      "VARCHAR(150)"),
            ("face_image",       "VARCHAR(500)"),
            ("aadhaar_pdf",      "VARCHAR(500)"), 
            ("account_name",     "VARCHAR(150)"),
            ("account_number",   "VARCHAR(30)"),
            ("ifsc",             "VARCHAR(15)"),
            ("pan",              "VARCHAR(12)"),
            ("project_name",     "VARCHAR(150)"),
        ]:
            try:
                cur.execute(f"ALTER TABLE employees ADD COLUMN {col} {defn}")
            except: pass
        for col, defn in [
            ("food_allowance",   "DECIMAL(10,2) DEFAULT 120.00"),
            ("food_before_time", "VARCHAR(5)    DEFAULT '08:00'"),
        ]:
            try:
                cur.execute(f"ALTER TABLE admin_settings ADD COLUMN {col} {defn}")
            except: pass

        try:
            cur.execute("ALTER TABLE employees ALTER COLUMN shift_hrs SET DEFAULT 9.0")
        except: pass
        try:
            cur.execute("UPDATE employees SET shift_hrs = 9.0 WHERE shift_hrs = 8.0")
        except: pass

        cur.execute("""INSERT IGNORE INTO admin_settings
            (id,pay_per_day,ot_pay_per_hr,food_allowance,food_before_time)
            VALUES (1,500.00,100.00,120.00,'08:00')""")
        db.commit(); db.close()
        print("Tables ready")
    except Exception as e:
        print(f"Startup warning: {e}")

def time_to_str(t):
    if t is None: return None
    if isinstance(t, str): return t[:5]
    if hasattr(t, "seconds"):
        total = int(t.total_seconds())
        h, rem = divmod(total, 3600)
        return f"{h:02d}:{rem//60:02d}"
    return str(t)[:5]

def time_before(t_str, limit_str):
    if not t_str: return False
    return t_str <= limit_str

# ── Models ───────────────────────────────────────────────────────────────────
class EmployeeCreate(BaseModel):
    full_name:       str
    father_name:     Optional[str] = None
    email:           str
    phone:           Optional[str] = None
    aadhaar_no:      str
    department:      str
    location:        Optional[str] = None
    source:          Optional[str] = None
    shift_hrs:       float = 9.0
    face_descriptor: Optional[List[float]] = None
    account_name:    Optional[str] = None
    account_number:  Optional[str] = None
    ifsc:            Optional[str] = None
    pan:             Optional[str] = None
    project_name:    Optional[str] = None

class EmployeeUpdate(BaseModel):
    source:          Optional[str] = None
    project_name:    Optional[str] = None
    account_name:    Optional[str] = None
    account_number:  Optional[str] = None
    ifsc:            Optional[str] = None
    pan:             Optional[str] = None

class AdminLogin(BaseModel):
    username: str
    password: str

class SourcePerson(BaseModel):
    name:           str
    account_name:   Optional[str] = None
    account_number: Optional[str] = None
    ifsc:           Optional[str] = None
    pan:            Optional[str] = None

class FaceDescriptorUpdate(BaseModel):
    face_descriptor: List[float]

class FaceImageUpdate(BaseModel):
    face_image: str   # Cloudinary URL

class AadhaarPdfUpdate(BaseModel):
    aadhaar_pdf: str  # Cloudinary URL

class ClockAction(BaseModel):
    emp_id: int

class ManualClockOut(BaseModel):
    log_out_time: str  # HH:MM

class SettingsUpdate(BaseModel):
    pay_per_day:      float
    ot_pay_per_hr:    float
    food_allowance:   float
    food_before_time: str

# ── EMPLOYEES ────────────────────────────────────────────────────────────────
@app.get("/employees")
def list_employees():
    db = get_db(); cur = db.cursor()
    cur.execute("""SELECT id, full_name, father_name, email, phone, aadhaar_no,
                          department, location, source, shift_hrs,
                          face_image, aadhaar_pdf,
                          account_name, account_number, ifsc, pan, project_name, created_at
                   FROM employees ORDER BY full_name""")
    result = list(cur.fetchall())
    for r in result:
        r["created_at"] = str(r["created_at"])
    db.close(); return result

@app.get("/employees/faces/all")
def get_all_faces():
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT id, full_name, face_descriptor FROM employees WHERE face_descriptor IS NOT NULL")
    result = []
    for row in cur.fetchall():
        fd = row["face_descriptor"]
        if isinstance(fd, str): fd = json.loads(fd)
        result.append({"id": row["id"], "full_name": row["full_name"], "face_descriptor": fd})
    db.close(); return result

@app.get("/employees/{emp_id}")
def get_employee(emp_id: int):
    db = get_db(); cur = db.cursor()
    cur.execute("""SELECT id, full_name, father_name, email, phone, aadhaar_no,
                          department, location, source, shift_hrs,
                          face_image, aadhaar_pdf,
                          account_name, account_number, ifsc, pan, project_name, created_at
                   FROM employees WHERE id=%s""", (emp_id,))
    row = cur.fetchone()
    if not row: db.close(); raise HTTPException(404, "Employee not found")
    row["created_at"] = str(row["created_at"])
    db.close(); return row

@app.post("/employees", status_code=201)
def create_employee(emp: EmployeeCreate):
    db = get_db(); cur = db.cursor()
    try:
        fd = json.dumps(emp.face_descriptor) if emp.face_descriptor else None
        cur.execute(
            """INSERT INTO employees
               (full_name, father_name, email, phone, aadhaar_no,
                department, location, source, shift_hrs, face_descriptor,
                account_name, account_number, ifsc, pan, project_name)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (emp.full_name, emp.father_name, emp.email, emp.phone,
             emp.aadhaar_no, emp.department, emp.location, emp.source, emp.shift_hrs, fd,
             emp.account_name, emp.account_number, emp.ifsc, emp.pan, emp.project_name))
        db.commit()
        new_id = cur.lastrowid
        # TiDB sometimes returns 0 for lastrowid — fetch the actual ID
        if not new_id:
            cur.execute("SELECT id FROM employees WHERE aadhaar_no=%s", (emp.aadhaar_no,))
            row = cur.fetchone()
            new_id = row["id"] if row else None
    except pymysql.IntegrityError as e:
        db.close()
        err = str(e)
        if "aadhaar_no" in err:
            raise HTTPException(409, "An employee with this Aadhaar number is already registered.")
        elif "email" in err:
            raise HTTPException(409, "An employee with this email is already registered.")
        else:
            raise HTTPException(409, "This employee already exists.")
    db.close()
    if not new_id:
        raise HTTPException(500, "Employee created but ID could not be retrieved. Please try again.")
    return {"message": "Employee registered", "id": new_id}

@app.put("/employees/{emp_id}")
def update_employee(emp_id: int, body: EmployeeUpdate):
    db = get_db(); cur = db.cursor()
    cur.execute(
        """UPDATE employees SET
           source=%s, project_name=%s, account_name=%s, account_number=%s, ifsc=%s, pan=%s
           WHERE id=%s""",
        (body.source, body.project_name, body.account_name, body.account_number, body.ifsc, body.pan, emp_id)
    )
    db.commit(); db.close()
    return {"message": "Employee details updated"}

@app.put("/employees/{emp_id}/face")
def update_face(emp_id: int, body: FaceDescriptorUpdate):
    db = get_db(); cur = db.cursor()
    cur.execute("UPDATE employees SET face_descriptor=%s WHERE id=%s",
                (json.dumps(body.face_descriptor), emp_id))
    db.commit(); db.close(); return {"message": "Face descriptor saved"}

@app.put("/employees/{emp_id}/face-image")
def update_face_image(emp_id: int, body: FaceImageUpdate):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT full_name FROM employees WHERE id=%s", (emp_id,))
    row = cur.fetchone()
    if not row:
        db.close()
        raise HTTPException(404, "Employee not found")
    full_name = row["full_name"].strip()
    
    face_image_val = body.face_image
    if face_image_val and not face_image_val.startswith("http"):
        try:
            filename = os.path.join(full_name, "face.jpg")
            face_image_val = save_base64_file(face_image_val, filename)
        except Exception as e:
            db.close()
            raise HTTPException(400, f"Failed to save face image: {e}")
    cur.execute("UPDATE employees SET face_image=%s WHERE id=%s",
                (face_image_val, emp_id))
    db.commit(); db.close(); return {"message": "Face image saved"}

@app.put("/employees/{emp_id}/aadhaar-pdf")
def update_aadhaar_pdf(emp_id: int, body: AadhaarPdfUpdate):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT full_name FROM employees WHERE id=%s", (emp_id,))
    row = cur.fetchone()
    if not row:
        db.close()
        raise HTTPException(404, "Employee not found")
    full_name = row["full_name"].strip()
    
    aadhaar_pdf_val = body.aadhaar_pdf
    if aadhaar_pdf_val and not aadhaar_pdf_val.startswith("http"):
        try:
            filename = os.path.join(full_name, "aadhaar.pdf")
            aadhaar_pdf_val = save_base64_file(aadhaar_pdf_val, filename)
        except Exception as e:
            db.close()
            raise HTTPException(400, f"Failed to save Aadhaar PDF: {e}")
    cur.execute("UPDATE employees SET aadhaar_pdf=%s WHERE id=%s",
                (aadhaar_pdf_val, emp_id))
    db.commit(); db.close(); return {"message": "Aadhaar PDF saved"}

class GetEmbeddingInput(BaseModel):
    image: str

@app.post("/get-embedding")
def get_embedding_route(data: GetEmbeddingInput):
    if face_app is not None:
        result = local_get_embedding(data.image)
        if result.get("success"):
            return result
        else:
            raise HTTPException(400, result.get("error", "Failed to extract face embedding locally"))
    else:
        result = hf_get_embedding(data.image)
        if result.get("success"):
            return result
        else:
            raise HTTPException(400, result.get("error", "Failed to extract face embedding from remote API"))

# Note: face_descriptor stores 512-d ArcFace embeddings (cosine similarity, threshold 0.4)

# ── ADMIN AUTH ──────────────────────────────────────────────────────────────
@app.post("/admin/login")
def admin_login(body: AdminLogin):
    admin_user = os.getenv("ADMIN_USERNAME", "admin")
    admin_pass = os.getenv("ADMIN_PASSWORD", "admin123")
    if body.username != admin_user or body.password != admin_pass:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    return {"username": body.username, "message": "Login successful"}

@app.post("/user/login")
def user_login(body: AdminLogin):
    user_name = os.getenv("USER_USERNAME", "user")
    user_pass = os.getenv("USER_PASSWORD", "user123")
    if body.username != user_name or body.password != user_pass:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    return {"username": body.username, "message": "Login successful"}

# ── SOURCE PERSONS ──────────────────────────────────────────────────────────
@app.get("/source-persons")
def list_source_persons():
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM source_persons ORDER BY name")
    result = list(cur.fetchall())
    for r in result: r["created_at"] = str(r["created_at"])
    db.close(); return result

@app.post("/source-persons", status_code=201)
def create_source_person(sp: SourcePerson):
    db = get_db(); cur = db.cursor()
    try:
        cur.execute("""INSERT INTO source_persons (name,account_name,account_number,ifsc,pan)
                       VALUES (%s,%s,%s,%s,%s)""",
                    (sp.name, sp.account_name, sp.account_number, sp.ifsc, sp.pan))
        db.commit(); new_id = cur.lastrowid
    except pymysql.IntegrityError as e:
        db.close(); raise HTTPException(409, "Source person already exists")
    db.close(); return {"message":"Source person added","id":new_id}

@app.delete("/source-persons/{sp_id}")
def delete_source_person(sp_id: int):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM source_persons WHERE id=%s", (sp_id,))
    db.commit(); db.close(); return {"message":"Removed"}

@app.delete("/employees/{emp_id}")
def delete_employee(emp_id: int):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM employees WHERE id=%s", (emp_id,))
    db.commit(); db.close(); return {"message": "Removed"}

# ── CLOCK IN / OUT ───────────────────────────────────────────────────────────
@app.post("/clock")
def clock(action: ClockAction):
    db = get_db(); cur = db.cursor()
    now_ist  = datetime.now(IST)
    today    = now_ist.date().isoformat()
    now_time = now_ist.strftime("%H:%M")
    cur.execute("SELECT * FROM employees WHERE id=%s", (action.emp_id,))
    row = cur.fetchone()
    if not row: db.close(); raise HTTPException(404, "Employee not found")
    emp = row; shift = float(emp["shift_hrs"])
    cur.execute("SELECT * FROM attendance WHERE emp_id=%s AND date=%s", (action.emp_id, today))
    att_row = cur.fetchone()
    if att_row is None:
        cur.execute("INSERT INTO attendance (emp_id,date,log_in,status) VALUES (%s,%s,%s,'on-duty')",
                    (action.emp_id, today, now_time))
        db.commit(); db.close()
        return {"action": "log_in", "time": now_time, "emp_name": emp["full_name"]}
    att = att_row
    ci = time_to_str(att["log_in"]); co = time_to_str(att["log_out"])
    if ci and not co:
        in_min  = int(ci[:2]) * 60 + int(ci[3:5])
        out_min = int(now_time[:2]) * 60 + int(now_time[3:5])
        total   = round((out_min - in_min) / 60, 2)
        ot      = round(max(0.0, total - 9.0), 2)
        if ot <= 0.5:
            ot = 0.0
        cur.execute("""UPDATE attendance SET log_out=%s,total_hrs=%s,ot_hrs=%s,status='present'
                       WHERE emp_id=%s AND date=%s""",
                    (now_time, total, ot, action.emp_id, today))
        db.commit(); db.close()
        return {"action": "log_out", "time": now_time,
                "total_hrs": total, "ot_hrs": ot, "emp_name": emp["full_name"]}
    db.close(); raise HTTPException(400, "Already completed attendance today")

@app.put("/attendance/{emp_id}/manual-clockout")
def manual_log_out(emp_id: int, date: str, body: ManualClockOut):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM attendance WHERE emp_id=%s AND date=%s", (emp_id, date))
    att = cur.fetchone()
    if not att: db.close(); raise HTTPException(404, "Attendance record not found")
    ci = time_to_str(att["log_in"])
    if not ci: db.close(); raise HTTPException(400, "No clock-in recorded")
    co = body.log_out_time
    in_min  = int(ci[:2])*60 + int(ci[3:5])
    out_min = int(co[:2])*60 + int(co[3:5])
    total   = round((out_min - in_min) / 60, 2)
    ot      = round(max(0.0, total - 9.0), 2)
    if ot <= 0.5:
        ot = 0.0
    cur.execute("""UPDATE attendance SET log_out=%s, total_hrs=%s, ot_hrs=%s, status='present'
                   WHERE emp_id=%s AND date=%s""", (co, total, ot, emp_id, date))
    db.commit(); db.close()
    return {"message": "Clock-out recorded", "log_out": co, "total_hrs": total, "ot_hrs": ot}

# ── ATTENDANCE ───────────────────────────────────────────────────────────────
@app.get("/attendance")
def get_attendance(emp_id: Optional[int]=None, month: Optional[str]=None, date_filter: Optional[str]=None):
    db = get_db(); cur = db.cursor()
    q = """SELECT a.*, e.full_name, e.email, e.phone, e.aadhaar_no,
                  e.department, e.location, e.source, e.shift_hrs, e.project_name
           FROM attendance a JOIN employees e ON a.emp_id=e.id WHERE 1=1"""
    p = []
    if emp_id:      q += " AND a.emp_id=%s";                       p.append(emp_id)
    if month:       q += " AND DATE_FORMAT(a.date,'%%Y-%%m')=%s";  p.append(month)
    if date_filter: q += " AND a.date=%s";                         p.append(date_filter)
    q += " ORDER BY a.date DESC, e.full_name"
    cur.execute(q, p)
    result = []
    for d in cur.fetchall():
        d["log_in"]  = time_to_str(d["log_in"])
        d["log_out"] = time_to_str(d["log_out"])
        d["date"]      = str(d["date"])
        result.append(d)
    db.close(); return result

@app.get("/attendance/today")
def get_today():
    return get_attendance(date_filter=datetime.now(IST).date().isoformat())

# ── DASHBOARD ────────────────────────────────────────────────────────────────
@app.get("/dashboard")
def dashboard():
    db = get_db(); cur = db.cursor()
    today = datetime.now(IST).date().isoformat()
    cur.execute("SELECT COUNT(*) as c FROM employees")
    total = cur.fetchone()["c"] or 0
    cur.execute("SELECT COUNT(*) as c FROM attendance WHERE date=%s AND status IN ('present','on-duty')", (today,))
    present = cur.fetchone()["c"] or 0
    cur.execute("""SELECT COUNT(*) as c FROM employees WHERE id NOT IN
        (SELECT emp_id FROM attendance WHERE date=%s AND status IN ('present','on-duty'))""", (today,))
    absent = cur.fetchone()["c"] or 0
    cur.execute("SELECT COALESCE(SUM(ot_hrs),0) as s FROM attendance WHERE date=%s", (today,))
    ot = float(cur.fetchone()["s"] or 0)
    cur.execute("""SELECT a.emp_id, a.log_in, e.full_name, e.project_name FROM attendance a
        JOIN employees e ON a.emp_id=e.id
        WHERE a.date=%s AND a.status='on-duty'""", (today,))
    on_duty = [{"emp_id": r["emp_id"], "log_in": time_to_str(r["log_in"]), "name": r["full_name"], "project_name": r["project_name"]}
               for r in cur.fetchall()]
    cur.execute("""SELECT id, full_name, department, project_name FROM employees WHERE id NOT IN
        (SELECT emp_id FROM attendance WHERE date=%s AND status IN ('present','on-duty'))""", (today,))
    absent_list = [{"id": r["id"], "name": r["full_name"], "dept": r["department"], "project_name": r["project_name"]}
                   for r in cur.fetchall()]
    db.close()
    return {"total_employees": total, "present": present, "absent": absent,
            "ot_hours": round(ot, 1), "on_duty": on_duty, "absent_employees": absent_list}

# ── ADMIN SETTINGS ───────────────────────────────────────────────────────────
@app.get("/settings")
def get_settings():
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT pay_per_day,ot_pay_per_hr,food_allowance,food_before_time FROM admin_settings WHERE id=1")
    row = cur.fetchone(); db.close()
    if not row:
        return {"pay_per_day": 500.0, "ot_pay_per_hr": 100.0,
                "food_allowance": 50.0, "food_before_time": "08:00"}
    return {"pay_per_day":      float(row["pay_per_day"]),
            "ot_pay_per_hr":    float(row["ot_pay_per_hr"]),
            "food_allowance":   float(row["food_allowance"]),
            "food_before_time": row["food_before_time"]}

@app.put("/settings")
def update_settings(s: SettingsUpdate):
    db = get_db(); cur = db.cursor()
    cur.execute("""UPDATE admin_settings
        SET pay_per_day=%s, ot_pay_per_hr=%s, food_allowance=%s,
            food_before_time=%s
        WHERE id=1""",
        (s.pay_per_day, s.ot_pay_per_hr, s.food_allowance, s.food_before_time))
    db.commit(); db.close(); return {"message": "Settings updated"}

# ── REPORTS EXCEL ────────────────────────────────────────────────────────────
def check_food_allowance(log_in_td, log_out_td):
    if not log_in_td or not log_out_td:
        return False
    in_sec = log_in_td.total_seconds()
    out_sec = log_out_td.total_seconds()
    return in_sec <= 8 * 3600 and out_sec >= 14 * 3600

def get_daily_ot_hours(log_in_td, log_out_td):
    if not log_in_td or not log_out_td:
        return 0.0
    in_sec = log_in_td.total_seconds()
    out_sec = log_out_td.total_seconds()
    total_hours = (out_sec - in_sec) / 3600.0
    ot_hours = total_hours - 9.0
    if ot_hours <= 0.5:
        return 0.0
    return ot_hours

@app.get("/reports/excel")
def report_excel(month: Optional[str]=None, date_filter: Optional[str]=None, emp_id: Optional[int]=None):
    if not month and not date_filter:
        raise HTTPException(400, "Provide month or date_filter")
        
    if month:
        year, month_num = map(int, month.split("-"))
    else:
        dt = datetime.strptime(date_filter, "%Y-%m-%d")
        year, month_num = dt.year, dt.month
        month = f"{year}-{month_num:02d}"
        
    import calendar
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    num_days = calendar.monthrange(year, month_num)[1]
    
    settings = get_settings()
    ppd = settings["pay_per_day"]
    otp = settings["ot_pay_per_hr"]
    food_amt = settings["food_allowance"]
    
    db = get_db(); cur = db.cursor()
    
    # 1. Fetch Source Persons for Banking details mapping
    cur.execute("SELECT name, account_name, account_number, ifsc, pan FROM source_persons")
    sources = cur.fetchall()
    source_map = {
        s["name"].lower().strip(): s for s in sources
    }
    
    # 2. Fetch Employees
    if emp_id:
        cur.execute("SELECT * FROM employees WHERE id=%s", (emp_id,))
    else:
        cur.execute("SELECT * FROM employees ORDER BY full_name")
    employees = cur.fetchall()
    
    # 3. Fetch Attendance
    start_date = f"{year}-{month_num:02d}-01"
    end_date = f"{year}-{month_num:02d}-{num_days:02d}"
    
    cur.execute("""
        SELECT a.*, e.full_name 
        FROM attendance a 
        JOIN employees e ON a.emp_id = e.id 
        WHERE a.date >= %s AND a.date <= %s
    """, (start_date, end_date))
    attendance_records = cur.fetchall()
    db.close()
    
    # Map attendance records
    att_map = {}
    active_date_strs = set()
    for r in attendance_records:
        e_id = r["emp_id"]
        d_str = r["date"].isoformat()
        if e_id not in att_map:
            att_map[e_id] = {}
        att_map[e_id][d_str] = r
        if r["status"] in ("present", "on-duty"):
            active_date_strs.add(d_str)
            
    # 4. Build column structure list dynamically
    cols = []
    cols.append({"header_name": "S.NO", "col_type": "sno"})
    cols.append({"header_name": "NAME", "col_type": "name"})
    cols.append({"header_name": "NUMBER", "col_type": "number"})
    cols.append({"header_name": "SOURCE", "col_type": "source"})
    cols.append({"header_name": "DESIGNATION", "col_type": "designation"})
    
    for d in range(1, num_days + 1):
        date_obj = date(year, month_num, d)
        date_iso = date_obj.isoformat()
        date_label = date_obj.strftime("%d-%b")
        cols.append({"header_name": date_label, "col_type": "date", "date_str": date_iso})
        
        # If activity exists on this date, show times, OT hours, and food allowance columns
        if date_iso in active_date_strs:
            cols.append({"header_name": "IN-TIME", "col_type": "in_time", "date_str": date_iso})
            cols.append({"header_name": "OUT-TIME", "col_type": "out_time", "date_str": date_iso})
            cols.append({"header_name": "OT-HOURS", "col_type": "ot_hours", "date_str": date_iso})
            cols.append({"header_name": "FOOD ALLOWANCE", "col_type": "food_allowance_daily", "date_str": date_iso})
            
    cols.append({"header_name": "WORKING DAYS", "col_type": "working_days"})
    cols.append({"header_name": "PER DAY", "col_type": "per_day"})
    cols.append({"header_name": "AMOUNT TO PAY", "col_type": "amount_to_pay"})
    cols.append({"header_name": "OT HOURS", "col_type": "total_ot_hours"})
    cols.append({"header_name": "FOOD ALLOWANCE", "col_type": "total_food_allowance"})
    cols.append({"header_name": "OT AMOUNT", "col_type": "total_ot_amount"})
    cols.append({"header_name": "TOTAL AMOUNT", "col_type": "total_amount"})
    cols.append({"header_name": "ACCOUNT NO", "col_type": "account_no"})
    cols.append({"header_name": "IFSC CODE", "col_type": "ifsc"})
    cols.append({"header_name": "PAN NO", "col_type": "pan"})
    cols.append({"header_name": "PAYEE NAME", "col_type": "payee_name"})
    
    # 5. Write Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance - {month}"
    
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=9)
    fill_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Write headers to Row 1
    for col_idx, col_info in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_info["header_name"])
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    ws.row_dimensions[1].height = 28
    
    # Write employee rows
    data_row_start = 2
    for idx, emp in enumerate(employees):
        r_num = data_row_start + idx
        emp_id = emp["id"]
        
        # Calculate payroll totals first
        total_present_days = 0
        total_ot_hours = 0.0
        total_food_allowance_days = 0
        
        for d in range(1, num_days + 1):
            date_str = f"{year}-{month_num:02d}-{d:02d}"
            record = att_map.get(emp_id, {}).get(date_str)
            if record:
                status = record["status"]
                is_worked = (status == "present" or status == "on-duty")
                if is_worked:
                    total_present_days += 1
                    
                log_in_td = record["log_in"]
                log_out_td = record["log_out"]
                
                daily_ot = get_daily_ot_hours(log_in_td, log_out_td)
                total_ot_hours += daily_ot
                
                if check_food_allowance(log_in_td, log_out_td):
                    total_food_allowance_days += 1
                    
        regular_pay = total_present_days * ppd
        ot_pay = total_ot_hours * otp
        food_allowance_pay = total_food_allowance_days * food_amt
        total_payout = regular_pay + ot_pay + food_allowance_pay
        
        # Banking details resolution (Sources details for source employees)
        emp_source = (emp.get("source") or "").lower().strip()
        acc_num = "—"
        ifsc_code = "—"
        pan_no = "—"
        payee_name = "—"
        
        if emp_source and emp_source != "ttipl" and emp_source in source_map:
            sp = source_map[emp_source]
            acc_num = sp.get("account_number") or "—"
            ifsc_code = sp.get("ifsc") or "—"
            pan_no = sp.get("pan") or "—"
            payee_name = sp.get("account_name") or sp.get("name") or "—"
        else:
            acc_num = emp.get("account_number") or "—"
            ifsc_code = emp.get("ifsc") or "—"
            pan_no = emp.get("pan") or "—"
            payee_name = emp.get("account_name") or emp.get("full_name") or "—"
            
        # Write columns
        for col_idx, col_info in enumerate(cols, 1):
            cell = ws.cell(row=r_num, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            c_type = col_info["col_type"]
            
            if c_type == "sno":
                cell.value = idx + 1
                cell.alignment = align_center
            elif c_type == "name":
                cell.value = emp["full_name"]
                cell.alignment = align_left
            elif c_type == "number":
                cell.value = emp.get("phone") or "—"
                cell.alignment = align_center
            elif c_type == "source":
                cell.value = emp.get("source") or "—"
                cell.alignment = align_left
            elif c_type == "designation":
                cell.value = emp.get("project_name") or "—"
                cell.alignment = align_left
            elif c_type == "date":
                date_iso = col_info["date_str"]
                record = att_map.get(emp_id, {}).get(date_iso)
                if record and record["status"] in ("present", "on-duty"):
                    cell.value = "P"
                else:
                    cell.value = "A"
                cell.alignment = align_center
            elif c_type == "in_time":
                date_iso = col_info["date_str"]
                record = att_map.get(emp_id, {}).get(date_iso)
                if record and record["status"] in ("present", "on-duty"):
                    cell.value = time_to_str(record["log_in"]) or "—"
                else:
                    cell.value = "—"
                cell.alignment = align_center
            elif c_type == "out_time":
                date_iso = col_info["date_str"]
                record = att_map.get(emp_id, {}).get(date_iso)
                if record and record["status"] in ("present", "on-duty"):
                    cell.value = time_to_str(record["log_out"]) or "—"
                else:
                    cell.value = "—"
                cell.alignment = align_center
            elif c_type == "ot_hours":
                date_iso = col_info["date_str"]
                record = att_map.get(emp_id, {}).get(date_iso)
                if record and record["status"] in ("present", "on-duty"):
                    daily_ot = get_daily_ot_hours(record["log_in"], record["log_out"])
                    cell.value = round(daily_ot, 1) if daily_ot > 0 else 0.0
                else:
                    cell.value = 0.0
                cell.alignment = align_center
                cell.number_format = '0.0'
            elif c_type == "food_allowance_daily":
                date_iso = col_info["date_str"]
                record = att_map.get(emp_id, {}).get(date_iso)
                if record and record["status"] in ("present", "on-duty"):
                    if check_food_allowance(record["log_in"], record["log_out"]):
                        cell.value = float(food_amt)
                    else:
                        cell.value = 0.0
                else:
                    cell.value = 0.0
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "working_days":
                cell.value = int(total_present_days)
                cell.alignment = align_center
            elif c_type == "per_day":
                cell.value = float(ppd)
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "amount_to_pay":
                cell.value = float(regular_pay)
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "total_ot_hours":
                cell.value = float(total_ot_hours)
                cell.alignment = align_center
                cell.number_format = '0.0'
            elif c_type == "total_food_allowance":
                cell.value = float(food_allowance_pay)
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "total_ot_amount":
                cell.value = float(ot_pay)
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "total_amount":
                cell.value = float(total_payout)
                cell.alignment = align_center
                cell.number_format = '₹#,##0.00'
            elif c_type == "account_no":
                cell.value = acc_num
                cell.alignment = align_center
            elif c_type == "ifsc":
                cell.value = ifsc_code
                cell.alignment = align_center
            elif c_type == "pan":
                cell.value = pan_no
                cell.alignment = align_center
            elif c_type == "payee_name":
                cell.value = payee_name
                cell.alignment = align_left if payee_name != "—" else align_center
                
        ws.row_dimensions[r_num].height = 20
        
    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 9)
        
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions[get_column_letter(len(cols))].width = 24
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    label = date_filter or month
    filename = f"attendance_report_{label}.xlsx"
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ── Serve React Frontend ───────────────────────────────────
dist_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"))

if os.path.exists(dist_path):
    # Serve static assets, icons, and face recognition models
    assets_dir = os.path.join(dist_path, "assets")
    if os.path.exists(assets_dir):
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")
        
    icons_dir = os.path.join(dist_path, "icons")
    if os.path.exists(icons_dir):
        app.mount("/icons", StaticFiles(directory=icons_dir), name="icons")
        
    models_dir = os.path.join(dist_path, "models")
    if os.path.exists(models_dir):
        app.mount("/models", StaticFiles(directory=models_dir), name="models")

    @app.get("/favicon.ico", include_in_schema=False)
    async def favicon():
        f = os.path.join(dist_path, "favicon.ico")
        return FileResponse(f) if os.path.exists(f) else FileResponse(os.path.join(dist_path, "index.html"))

    @app.get("/", include_in_schema=False)
    async def serve_root():
        return FileResponse(os.path.join(dist_path, "index.html"))

    @app.get("/manifest.json")
    async def serve_manifest():
        return FileResponse(
            os.path.join(dist_path, "manifest.json"),
            headers={"Content-Type": "application/manifest+json"}
        )

    # Catch-all: return index.html for React Router paths
    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_frontend(full_path: str):
        return FileResponse(os.path.join(dist_path, "index.html"))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
